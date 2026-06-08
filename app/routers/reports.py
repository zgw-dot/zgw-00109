from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
from datetime import datetime
import json
from io import BytesIO

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from ..database import get_db
from .. import models, schemas

router = APIRouter(
    prefix="/api/reports",
    tags=["报告导出"],
    responses={404: {"description": "未找到"}}
)


def _get_package_data(db: Session, package_no: str):
    package = db.query(models.TaskPackage).filter(
        models.TaskPackage.package_no == package_no
    ).first()
    if not package:
        return None, None, None, None

    template = db.query(models.InspectionTemplate).filter(
        models.InspectionTemplate.id == package.template_id
    ).first()

    readings = db.query(models.Reading).filter(
        models.Reading.task_package_id == package.id
    ).order_by(models.Reading.device_code, models.Reading.item_name).all()

    conflicts = db.query(models.Conflict).filter(
        models.Conflict.task_package_id == package.id
    ).order_by(models.Conflict.created_at.desc()).all()

    return package, template, readings, conflicts


@router.get("/{package_no}/json")
def export_report_json(package_no: str, db: Session = Depends(get_db)):
    package, template, readings, conflicts = _get_package_data(db, package_no)
    if not package:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"任务包编号 '{package_no}' 不存在"
        )

    report_data = {
        "package_info": {
            "package_no": package.package_no,
            "status": package.status,
            "operator": package.operator,
            "created_at": package.created_at.isoformat() if package.created_at else None,
            "issued_at": package.issued_at.isoformat() if package.issued_at else None,
            "synced_at": package.synced_at.isoformat() if package.synced_at else None,
            "closed_at": package.closed_at.isoformat() if package.closed_at else None,
        },
        "template_info": {
            "id": template.id,
            "name": template.name,
            "description": template.description,
            "check_items_count": len(template.check_items) if template else 0
        } if template else None,
        "readings": [
            {
                "id": r.id,
                "device_code": r.device_code,
                "item_name": r.item_name,
                "reading_value": r.reading_value,
                "collected_at": r.collected_at.isoformat(),
                "uploaded_at": r.uploaded_at.isoformat(),
                "source_type": r.source_type
            } for r in readings
        ],
        "conflicts": [
            {
                "id": c.id,
                "device_code": c.device_code,
                "item_name": c.item_name,
                "existing_value": c.existing_value,
                "new_value": c.new_value,
                "status": c.status,
                "resolution_note": c.resolution_note,
                "created_at": c.created_at.isoformat(),
                "resolved_at": c.resolved_at.isoformat() if c.resolved_at else None,
                "resolved_by": c.resolved_by
            } for c in conflicts
        ],
        "summary": {
            "total_readings": len(readings),
            "total_conflicts": len(conflicts),
            "open_conflicts": len([c for c in conflicts if c.status == "open"]),
            "resolved_conflicts": len([c for c in conflicts if c.status == "resolved"])
        },
        "exported_at": datetime.now().isoformat()
    }

    json_str = json.dumps(report_data, ensure_ascii=False, indent=2)
    bytes_io = BytesIO(json_str.encode("utf-8"))

    return StreamingResponse(
        bytes_io,
        media_type="application/json",
        headers={
            "Content-Disposition": f"attachment; filename=report_{package_no}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        }
    )


@router.get("/{package_no}/excel")
def export_report_excel(package_no: str, db: Session = Depends(get_db)):
    if not HAS_OPENPYXL:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel 导出功能需要 openpyxl 库，请先安装: pip install openpyxl"
        )

    package, template, readings, conflicts = _get_package_data(db, package_no)
    if not package:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"任务包编号 '{package_no}' 不存在"
        )

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "任务包信息"

    headers = ["字段", "值"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    info_data = [
        ["任务包编号", package.package_no],
        ["状态", package.status],
        ["操作人员", package.operator or "-"],
        ["创建时间", package.created_at.strftime("%Y-%m-%d %H:%M:%S") if package.created_at else "-"],
        ["发放时间", package.issued_at.strftime("%Y-%m-%d %H:%M:%S") if package.issued_at else "-"],
        ["同步时间", package.synced_at.strftime("%Y-%m-%d %H:%M:%S") if package.synced_at else "-"],
        ["关闭时间", package.closed_at.strftime("%Y-%m-%d %H:%M:%S") if package.closed_at else "-"],
        ["模板名称", template.name if template else "-"],
        ["模板描述", template.description if template else "-"],
        ["检查项数量", len(template.check_items) if template else 0],
        ["读数总数", len(readings)],
        ["冲突总数", len(conflicts)],
        ["未解决冲突", len([c for c in conflicts if c.status == "open"])],
        ["已解决冲突", len([c for c in conflicts if c.status == "resolved"])],
    ]

    for row, data in enumerate(info_data, 2):
        for col, value in enumerate(data, 1):
            ws1.cell(row=row, column=col, value=value)

    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 50

    ws2 = wb.create_sheet("读数记录")
    headers = ["ID", "设备编号", "检查项", "读数", "采集时间", "上传时间", "来源"]
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for row, reading in enumerate(readings, 2):
        ws2.cell(row=row, column=1, value=reading.id)
        ws2.cell(row=row, column=2, value=reading.device_code)
        ws2.cell(row=row, column=3, value=reading.item_name)
        ws2.cell(row=row, column=4, value=reading.reading_value)
        ws2.cell(row=row, column=5, value=reading.collected_at.strftime("%Y-%m-%d %H:%M:%S"))
        ws2.cell(row=row, column=6, value=reading.uploaded_at.strftime("%Y-%m-%d %H:%M:%S"))
        ws2.cell(row=row, column=7, value=reading.source_type)

    for col in range(1, 8):
        ws2.column_dimensions[chr(64 + col)].width = 20

    ws3 = wb.create_sheet("冲突记录")
    headers = ["ID", "设备编号", "检查项", "已有读数", "新读数", "状态", "解决说明", "创建时间", "解决时间", "解决人"]
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FFC000" if len(conflicts) > 0 else "4472C4", end_color="FFC000" if len(conflicts) > 0 else "4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for row, conflict in enumerate(conflicts, 2):
        ws3.cell(row=row, column=1, value=conflict.id)
        ws3.cell(row=row, column=2, value=conflict.device_code)
        ws3.cell(row=row, column=3, value=conflict.item_name)
        ws3.cell(row=row, column=4, value=conflict.existing_value)
        ws3.cell(row=row, column=5, value=conflict.new_value)
        ws3.cell(row=row, column=6, value=conflict.status)
        ws3.cell(row=row, column=7, value=conflict.resolution_note or "-")
        ws3.cell(row=row, column=8, value=conflict.created_at.strftime("%Y-%m-%d %H:%M:%S"))
        ws3.cell(row=row, column=9, value=conflict.resolved_at.strftime("%Y-%m-%d %H:%M:%S") if conflict.resolved_at else "-")
        ws3.cell(row=row, column=10, value=conflict.resolved_by or "-")

    for col in range(1, 11):
        ws3.column_dimensions[chr(64 + col)].width = 18

    bytes_io = BytesIO()
    wb.save(bytes_io)
    bytes_io.seek(0)

    filename = f"report_{package_no}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        bytes_io,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
